import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from api.utils.debt import get_debt_list


def create_debt_list_file(
    currency: str, branch: str, limit=1000, filename="DebtList.xlsx"
):
    wb = Workbook()
    ws = wb.active

    COLUMN_LIST = ["ИД контрагента", "Контрагент", "Валюта", "Долг"]

    # Set column widths (optional)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20

    ws.column_dimensions["A"].height = 20
    ws.column_dimensions["B"].height = 20
    ws.column_dimensions["C"].height = 20
    ws.column_dimensions["D"].height = 20

    # Define custom styles
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(
        start_color="C4C4C4", end_color="C4C4C4", fill_type="solid"
    )
    alignment_center = Alignment(horizontal="center")

    for col_num, header in enumerate(COLUMN_LIST, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    debt_list = get_debt_list(branch, currency, limit)["customers"]
    for row_num, debt_item in enumerate(debt_list, start=2):
        if debt_item["amount"] == 0:
            continue

        debt_item["amount"] = "{:,.2f}".format(debt_item["amount"]).replace(",", " ")
        row = [
            debt_item["id"],
            debt_item["name"],
            debt_item["currency"],
            debt_item["amount"],
        ]
        ws.append(row)

        for col_num in range(1, 5):
            if col_num != 2:
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = alignment_center

    wb.save(filename)
    file_path = os.path.abspath(filename)
    return file_path
